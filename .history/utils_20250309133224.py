import os
import sys
import pandas as pd
import openpyxl
import subprocess
import platform
import shutil
from pathlib import Path
from openpyxl.utils import get_column_letter
from config import get_csv_encoding, get_date_format, get_input_dir, get_output_dir

def setup_directories():
    """
    必要なフォルダとファイルを作成する
    """
    # 現在のexeファイルのディレクトリを取得
    current_dir = Path(os.path.dirname(os.path.abspath(__file__)))
    
    # 必要なフォルダを作成
    folders = ['input', 'output', 'templates']
    for folder in folders:
        folder_path = current_dir / folder
        if not folder_path.exists():
            folder_path.mkdir(parents=True)

    # テンプレートファイルをコピー
    template_path = current_dir / 'templates' / '勤怠表雛形_2025年版.xlsx'
    if not template_path.exists():
        if getattr(sys, '_MEIPASS', None):
            # PyInstallerで実行している場合
            bundled_template = Path(sys._MEIPASS) / 'templates' / '勤怠表雛形_2025年版.xlsx'
            if bundled_template.exists():
                os.makedirs(template_path.parent, exist_ok=True)
                shutil.copy2(str(bundled_template), str(template_path))
            else:
                # テンプレートが見つからない場合は空のテンプレートを作成
                wb = openpyxl.Workbook()
                sheet = wb.active
                sheet.title = "勤務表"
                wb.save(template_path)


def open_folder(path):
    """
    指定されたフォルダをOSのデフォルトファイルエクスプローラで開く
    """
    if not os.path.exists(path):
        os.makedirs(path, exist_ok=True)
        
    try:
        if platform.system() == "Windows":
            os.startfile(path)
        elif platform.system() == "Darwin":  # macOS
            subprocess.run(['open', path], check=False)
        else:  # Linux
            subprocess.run(['xdg-open', path], check=False)
        return True
    except Exception:
        return False


def read_csv(csv_path):
    """
    CSVファイルを読み込み、DataFrameとして返す
    """
    df = pd.read_csv(csv_path, encoding=get_csv_encoding())
    df["日付"] = pd.to_datetime(df["日付"], format=get_date_format())
    return df

def process_data(df):
    """
    勤怠データを整形する
    """
    # 必要なカラムの選択
    columns_to_keep = ["日付", "始業時刻", "終業時刻", "総勤務時間", "法定内残業", "時間外労働", "深夜労働", "勤怠種別"]
    df_filtered = df[columns_to_keep].copy()
    
    # 不要データの削除（未入力、休日）
    # 未入力・休日データも残すように変更
    # df_filtered = df_filtered[~df_filtered["勤怠種別"].isin(["未入力", "所定休日", "法定休日"])]
    
    # 時間を小数時間に変換
    def time_to_hours(time_str):
        if isinstance(time_str, str) and ":" in time_str:
            h, m = map(int, time_str.split(":"))
            return h + m / 60  # 分を時間に変換
        return 0

    df_filtered["総勤務時間"] = df_filtered["総勤務時間"].apply(time_to_hours)
    
    return df_filtered

def write_to_excel(template_path, output_path, df, csv_filename):
    """
    ひな型Excelに勤怠データを書き込む
    """
    import openpyxl
    wb = openpyxl.load_workbook(template_path)
    sheet = wb["勤務表"]

    # G6セルにCSVファイル名から取得した従業員名を記載
    import re
    name_match = re.search(r'勤怠詳細_(.+?)_\d{4}_\d{2}', os.path.basename(csv_filename))
    employee_name = name_match.group(1) if name_match else "不明"
    sheet["G6"] = employee_name

    # H5セルの月を取得し、それに基づいてA列の日付を設定
    month_value = df["日付"].dt.month.iloc[0]
    year_value = df["日付"].dt.year.iloc[0]
    sheet["H5"] = month_value
    sheet["F5"] = year_value

    for index, row in enumerate(df.itertuples(), start=11):  # C列から開始
        day_value = index - 10  # A11に1日から入力
        sheet[f"A{index}"] = f"=DATE({year_value},{month_value},{day_value})"
        sheet[f"C{index}"] = row.始業時刻
        sheet[f"D{index}"] = row.終業時刻
        sheet[f"E{index}"] = "1:00" if row.勤怠種別 not in ["未入力", "所定休日", "法定休日"] else ""  # 休憩時間
        sheet[f"F{index}"] = row.総勤務時間  

    wb.save(output_path)
    print(f"✅ Excelファイルを保存しました: {output_path}")
